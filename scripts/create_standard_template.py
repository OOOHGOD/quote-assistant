"""生成项目内置的标准报价单模板。

这个脚本只用于生成示例/默认模板；真实业务模板仍建议通过 `import-template` 体检后启用。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = ROOT / "templates"
OUTPUT_PATH = TEMPLATE_DIR / "standard-quotation-template.xlsm"


def apply_table_style(ws) -> None:
    """给报价单工作表设置基础样式、列宽和数字格式。"""
    thin = Side(style="thin", color="B7C3D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    section_fill = PatternFill("solid", fgColor="F4F7FB")
    bold = Font(bold=True, color="1F1F1F")
    center = Alignment(horizontal="center", vertical="center")

    for row in range(1, 33):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if row in {1, 7, 30, 31, 32}:
                cell.fill = header_fill if row == 7 else section_fill
            if row == 7:
                cell.font = bold
                cell.alignment = center
            elif row in {1, 2, 3, 4, 5, 6, 30, 31, 32} and col == 1:
                cell.font = bold

    for cell in ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A30", "A31", "A32"]:
        ws[cell].font = bold

    for row in range(8, 28):
        for col in range(1, 13):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center")

    for address in ["A7", "B7", "C7", "D7", "E7", "F7", "G7", "H7", "I7", "J7", "K7", "L7"]:
        ws[address].alignment = center

    for address in ["I8", "I9", "I10", "I11", "I12", "I13", "I14", "I15", "I16", "I17", "I18", "I19", "I20", "I21", "I22", "I23", "I24", "I25", "I26", "I27", "J8", "J9", "J10", "J11", "J12", "J13", "J14", "J15", "J16", "J17", "J18", "J19", "J20", "J21", "J22", "J23", "J24", "J25", "J26", "J27", "B30", "B31", "B32"]:
        ws[address].number_format = '#,##0.00'

    ws["B5"].number_format = "yyyy-mm-dd"


def build_template() -> Path:
    """创建标准报价单模板并保存到 templates 目录。"""
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "报价单"

    headers = [
        ("A1", "报价编号"), ("A2", "供应商"), ("A3", "客户"),
        ("A4", "项目"), ("A5", "报价日期"), ("A6", "币种"),
    ]
    for cell, value in headers:
        ws[cell] = value

    table_headers = [
        "序号", "产品编码", "产品名称", "规格尺寸", "材质", "颜色",
        "单位", "数量", "单价", "金额", "房间/区域", "备注",
    ]
    for idx, label in enumerate(table_headers, start=1):
        ws.cell(row=7, column=idx, value=label)

    totals = [("A30", "小计"), ("A31", "税额"), ("A32", "总计")]
    for cell, value in totals:
        ws[cell] = value

    ws.freeze_panes = "A8"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 24

    apply_table_style(ws)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = build_template()
    print(path)
